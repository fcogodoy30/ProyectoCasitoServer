from django.shortcuts import render, redirect
from django.contrib.auth.models import User
from django.contrib.auth import login, logout, authenticate
from django.db import IntegrityError
from .models import CasinoColacion, Estado, Opciones, Usuarios, TipoUsuario, Programacion
from django.contrib import messages 
from datetime import datetime, timedelta
from django.utils import timezone
from collections import defaultdict
from django.shortcuts import get_object_or_404
from django.db import transaction
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
import json
import pytz

from django.contrib.auth.decorators import user_passes_test

def check_user_profiles(user, allowed_profiles):
    return user.is_authenticated and user.usuarios.tipo_usuario.tipo in allowed_profiles

def check_soporte_or_admin(user):
    return check_user_profiles(user, ['Soporte', 'Empresa'])

def check_soporte(user):
    return check_user_profiles(user, ['Soporte'])

def check_admin(user):
    return check_user_profiles(user, ['Empresa'])


#login
def home(request):
    return render(request,'login/login.html')

@login_required
def primeringreso(request):
    if request.method == 'GET':
        return render(request,'primeringreso.html')
    else:
        id_user = request.POST.get('id_user')
        password = request.POST.get('password1')
        usuario = get_object_or_404(User, id=id_user)
        if password:
            usuario.set_password(password)  # Actualizamos el Password
        usuario.save()
        logout(request)
        request.session.flush()
        return render(request,'login/login.html',{
                          'msg': 'Contraseña actualizada, ingrese nuevamente '})
        
  
# PRINCIPAL
@login_required
def principal(request):
    user_data = request.session.get('user_data', {})
    current_date = datetime.now().date()  # Obtén la fecha actual
    return render(request, 'principal.html', {
        'user_data': user_data,
        'current_date': current_date,
    })
    
#CERRAR SESION
@login_required
def cerrarsession(request):
    logout(request)
    request.session.flush()
    return redirect('home')

#INICIO SESION
def iniciosession(request):
    if request.method == 'GET':
        return render(request,'login/login.html')
    else:
        user = authenticate(request, username=request.POST['username'] , password=request.POST['password'])
            
        if user is None:
              
            return render(request,'login/login.html',{
                          'error': 'Rut o Contraseña incorrecto'})
        else:
            activo = Usuarios.objects.filter(rut = request.POST['username'], activo = 1)
            if activo:
                ulticonex = user.last_login
            
                if ulticonex is not None:
                    login(request, user)
                           
                    request.session['user_data'] = {
                        'id' : user.id,
                        'username': user.username,
                        'nombre': user.first_name,
                        'apellido': user.last_name,
                        
                    }
                    return redirect('principal')
                else:
                    login(request, user)
                    return render(request, 'primeringreso.html', {
                        'id': user.id
                    })
            else:
                return render(request,'login/login.html',{
                          'error': 'Rut Inactivo'})

#Editar Usuarios
@login_required
@user_passes_test(check_soporte_or_admin)
def editusuario(request, id):
    if request.method == 'GET':
        #consulta = request.GET.get('q')
        # Guardamos el Perfil que esta solicitando el dato
        usuario_actual = Usuarios.objects.get(id_user=request.user)
        tipo_usuario_actual = usuario_actual.tipo_usuario
        
        if tipo_usuario_actual.tipo == 'Soporte':
            usuario = Usuarios.objects.filter(id_user_id=id).order_by('tipo_usuario')
            tipousuario = TipoUsuario.objects.all().order_by('id')
        else:
            usuario = Usuarios.objects.filter(id_user_id=id)
            tipousuario = TipoUsuario.objects.filter(id__in=[1, 2]).order_by('id')
    
        
        context = {
            'usuarios': usuario,
            #'query': consulta, #Resultado para buscador comentado
            'tipousuario': tipousuario,
            'id_user' : id
        }
        return render(request, 'adminCliente/edit_usuarios.html', context)
    else:
        
        id_user = request.POST.get('id_user')
        nombre = request.POST.get('first_name')
        apellido = request.POST.get('last_name')
        password = request.POST.get('password1')
        tiposus = request.POST.get('tipousuario')

        # Aqui se guarda la tabla User para Login
        user = get_object_or_404(User, id=id_user)
        user.first_name = nombre
        user.last_name = apellido
        if password:
            user.set_password(password)  # Actualizamos el Password
        user.save()
        
        tipo_usuario = get_object_or_404(TipoUsuario, pk=tiposus)
                
        # Aqui guardamos los datos en la tabla del usuario
        usuario = get_object_or_404(Usuarios, id_user=id_user)
        usuario.nombre = nombre
        usuario.apellido = apellido
        usuario.tipo_usuario = tipo_usuario
        usuario.save()
        
        messages.success(request, "Registro Actualizado.")
        return redirect('usuarioslistas')

# Dentro tenemos el guardado del Usuario
@login_required
@user_passes_test(check_soporte_or_admin)
def usuarios(request):
    if request.method == 'GET':
        return redirect('usuarioslistas')
    else:
        if request.POST['password1'] == request.POST['password2']:
            try:
                with transaction.atomic(): # Lo utilizaremos para asegurar de haber alguna falla no haga ninguna ejecucion a las tablas
                    rut = request.POST.get('username')
                    password = request.POST.get('password1')
                    first_name = request.POST.get('first_name')
                    last_name = request.POST.get('last_name')
                    tipoUsu = request.POST.get('tipousuario')
                    
                    user = User.objects.create_user(username=rut, password=password, first_name=first_name, last_name=last_name)
                    user.save()
                    # Obtener la instancia de TipoUsuario
                    tipo_usuario = get_object_or_404(TipoUsuario, pk=tipoUsu)
                                            
                    usuario = Usuarios.objects.create(rut = rut, nombre = first_name, apellido=last_name,  id_user=user, tipo_usuario = tipo_usuario)
                    usuario.save()
                    messages.success(request, f'El rut {rut} se ha registrado con exito')
                    return redirect ('usuarioslistas')
            except IntegrityError as e:
                messages.warning(request, f'Rut {rut} ya se ecuentra registrado')
                return redirect('usuarioslistas')
            
            except Exception as e:
                messages.error(request, f"Ocurrió un error al guardar el usuario: {e}")
                return redirect('usuarioslistas')                
        else: 
            messages.warning(request, 'Contraseña no coinciden')
            return redirect('usuarioslistas')

# lista de usuarios
@login_required
@user_passes_test(check_soporte_or_admin)
def usuarioslistas(request):
    #consulta = request.GET.get('q')
        # Guardamos el Perfil que esta solicitando el dato
    usuario_actual = Usuarios.objects.get(id_user=request.user)
    tipo_usuario_actual = usuario_actual.tipo_usuario
    
    if tipo_usuario_actual.tipo == 'Soporte':
        usuarios = Usuarios.objects.all().order_by('tipo_usuario')
        tipousuario = TipoUsuario.objects.all().order_by('id')
    else:
        usuarios = Usuarios.objects.filter(tipo_usuario__in=[1, 2]).order_by('tipo_usuario')
        tipousuario = TipoUsuario.objects.filter(id__in=[1, 2]).order_by('id')
             
    context = {
        'usuarios': usuarios,
        #'query': consulta, #Resultado para buscador comentado
        'tipousuario': tipousuario,
        'messages': messages.get_messages(request),
    }
    return render(request, 'adminCliente/usuarios.html', context)

#EDITAR MENU
@login_required
@user_passes_test(check_soporte)
def editamenu(request, id):
    if request.method == 'GET':
        #consulta = request.GET.get('q')
        #if consulta:
        #    menu = CasinoColacion.objects.filter(titulo__icontains=consulta).order_by('fecha_servicio')
        
        menu = CasinoColacion.objects.filter(id=id).order_by('fecha_servicio', 'id_opciones__opciones')
            
        estado = Estado.objects.all().order_by('id')
        opcion = Opciones.objects.all().order_by('id')
        context = {
            'menus': menu,
            #'query': consulta,
            'estados': estado,
            'opciones': opcion,
            'id_menu': id
        }
        return render(request, 'admin/edit_agregarmenu.html', context)
    else:
        id_menu = request.POST.get('id_menu')
        titulo = request.POST.get('titulo')
        fecha_servicio = request.POST.get('fechaServicio')
        descripcion = request.POST.get('descripcion')
        # obtengo el objeto
        menu_item = get_object_or_404(CasinoColacion, id=id_menu)
        # Actualizamos Campos
        menu_item.titulo = titulo
        menu_item.fecha_servicio = fecha_servicio
        menu_item.descripcion = descripcion
        menu_item.save() #guarda
        messages.success(request, "Registro Actualizado.")
        return redirect('menu_lista')
        
#AGREGAR MENU
@login_required
@user_passes_test(check_soporte)
def agregarmenu(request):
    if request.method == 'GET':
        return redirect('menu_lista')
    else:  
        try:
            with transaction.atomic():
                titulo = request.POST.get('titulo')
                fechaSer = request.POST.get('fechaServicio')
                estado = request.POST.get('estado')
                opcion = request.POST.get('opcion')
                desc = request.POST.get('descripcion')
                
                # Validar si el título ya existe
                if CasinoColacion.objects.filter(titulo=titulo, fecha_servicio=fechaSer).exists():
                    messages.error(request, "El título ya existe.")
                    return redirect('menu_lista')

                # Validar si hay opciones repetidas con las mismas fechas
                if CasinoColacion.objects.filter(id_opciones_id=opcion, fecha_servicio=fechaSer).exists():
                    messages.error(request, "Ya existe una opción con la misma fecha.")
                    return redirect('menu_lista')            
                registro = CasinoColacion(titulo = titulo ,descripcion = desc, fecha_servicio = fechaSer, 
                                        id_estado = estado, id_opciones_id = opcion)
                registro.save()
                messages.success(request, "Menu Guardado con exito.")
                return redirect('menu_lista')
        except IntegrityError as e:
            messages.error(request, "Error al agregar el menu.")

#ELIMINAR MENU
@login_required
@user_passes_test(check_soporte)
def eliminarMenu(request, id):
    menu = get_object_or_404(CasinoColacion, id=id)
        
    try:
        # Confirmación de eliminación
        menu.delete()
        messages.success(request, "Menú eliminado correctamente.")
        return redirect('menu_lista')
    except Exception as e:
               messages.error(request, f"Error al eliminar el menú: {e}")
    return redirect('menu_lista')  # Redirige a la lista de menús después de cualquier acción
        
#LISTA DEL MENU    
@login_required
@user_passes_test(check_soporte)
def menu_lista(request):
    #consulta = request.GET.get('q')
    #if consulta:
    #    menu = CasinoColacion.objects.filter(titulo__icontains=consulta).order_by('fecha_servicio', 'id_opciones__opciones')
    menu = CasinoColacion.objects.all().order_by('fecha_servicio', 'id_opciones__opciones')
    
    estado = Estado.objects.all().order_by('id')
    opcion = Opciones.objects.all().order_by('id')
    context = {
        'menus': menu,
        #'query': consulta,
        'estados': estado,
        'opciones': opcion
    }
    return render(request, 'admin/agregarmenu.html', context)

#Cambia el estado del usuario
@csrf_exempt
def cambiar_estado_usuario(request):
    if request.method == 'POST':
        
        usuario_id = request.POST.get('usuario_id')
        activo = request.POST.get('activo')
        usuario = Usuarios.objects.get(id=usuario_id)
        usuario.activo = bool(int(activo))
        usuario.save()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False})

@csrf_exempt
def cambiar_estado_menu(request):
    if request.method == 'POST':   
        id = request.POST.get('menu_Id')
        estado = request.POST.get('activo')
        menu = CasinoColacion.objects.get(id=id)
        menu.id_estado = estado
        menu.save()
        return JsonResponse({'success': True})    
    return JsonResponse({'success': False})

# CREACION METODO SELECCION SEMANA ACTIVA
def diaDeSemana():
    # Obtener la fecha actual
    fechaActual = datetime.now().date()
    # Calcular el inicio de la semana actual (lunes)
    inicioSem = fechaActual - timedelta(days=fechaActual.weekday())
    
    # Calcular el inicio y fin de la siguiente semana (lunes a viernes)
    inicioSemSiguiente = inicioSem + timedelta(days=7)
    finSemSiguiente = inicioSemSiguiente + timedelta(days=4)

    return inicioSemSiguiente, finSemSiguiente

#Programar menu para semana 
@login_required
def programarmenu(request):
        user_id = request.user.id
        iniSem, finSem = diaDeSemana()
        # Consultar las programaciones del usuario para la semana
        programaciones_usuario = Programacion.objects.filter(
            usuario__id_user=user_id,
            fecha_servicio__range=[iniSem, finSem]
        )

        # Si alguna fecha está activa, redirige a una página de error o muestra un mensaje
        if programaciones_usuario:
            semana_activa = Programacion.objects.filter(
            usuario__id_user=user_id,
            fecha_servicio__range=[iniSem, finSem]
            ).first()
            mensaje = f"Tu menu del dia {semana_activa.fecha_servicio} al {semana_activa.fecha_servicio + timedelta(days=4)} ya fue seleccionado."
            #return render(request, 'error.html', {'message': mensaje})
            messages.success(request, mensaje)
            return redirect('principal')
        
        # Si no hay fechas activas, continúa con la lógica normal
        
        # Obtener la programación de CasinoColacion para la semana
        programacion = CasinoColacion.objects.filter(fecha_servicio__range=[iniSem, finSem], id_estado=1).order_by('fecha_servicio', 'id_opciones_id')
        
        # Agrupar por fecha
        programacion_dict = defaultdict(list)
        for registro in programacion:
            programacion_dict[registro.fecha_servicio].append(registro)
            
        # Ordenar la programación por fecha
        programacion_ordenada = sorted(programacion_dict.items())
        
        TipoUsuario = Usuarios.objects.get(id_user=user_id)
        
        if TipoUsuario.tipo_usuario_id == 1:
            return render(request, 'usuario/programarmenu_emp.html', {'programacion_ordenada': programacion_ordenada})
        else:
            return render(request, 'usuario/programarmenu.html', {'programacion_ordenada': programacion_ordenada})

@csrf_exempt  # Desactiva la verificación CSRF para facilitar el desarrollo
def guardar_selecciones(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        usuario = Usuarios.objects.get(id_user_id=request.user.id)  # Asegúrate de que el usuario esté autenticado
        for item in data:
            fecha_servicio = datetime.strptime(item['fecha_servicio'], '%Y-%m-%d').date()
            casino_colacion = CasinoColacion.objects.get(id=item['opcion_id'])
            cantidad = item['cant']
                
            now = timezone.now()
            santiago_tz = pytz.timezone('America/Santiago')
            now_santiago = now.astimezone(santiago_tz)

                        
            Programacion.objects.create(
                usuario=usuario,
                menu_id=casino_colacion.id,
                nom_menu=casino_colacion,
                fecha_servicio=fecha_servicio,
                cantidad_almuerzo=cantidad,
                fecha_seleccion=now_santiago,
                impreso=0,
                origen = 'nube',
                _syncing=casino_colacion.id_opciones.id #Ocuparemos la variable para tener el ID de Opcion menu
            )
            
            
        return JsonResponse({ 'status': 'success' })  # Redirige a la página principal
    return JsonResponse({'status': 'fail'}, status=400)

#control y descarga
@login_required
@user_passes_test(check_soporte_or_admin)
def control_descarga(request):
    return render(request, 'admin/control_descarga.html')


from django_filters.views import FilterView
from django_tables2.views import SingleTableMixin
from .models import Programacion
from .filters import ProgramacionFilter
from .tables import ProgramacionTable
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import openpyxl  # Añadir esta línea para importar openpyxl
from openpyxl.utils import get_column_letter 



class ProgramacionListView(SingleTableMixin, FilterView):
    model = Programacion
    table_class = ProgramacionTable
    template_name = 'admin/control_descarga.html'
    filterset_class = ProgramacionFilter

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter'] = ProgramacionFilter(self.request.GET, queryset=self.get_queryset())
        return context

@login_required
@user_passes_test(check_soporte_or_admin)
def export_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="programacion.xlsx"'

    # Aplicar los mismos filtros de la vista ProgramacionListView
    filter = ProgramacionFilter(request.GET, queryset=Programacion.objects.all())
    programaciones = filter.qs

    # Crear un libro de trabajo y una hoja de trabajo
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Programacion Menu Usuarios"

    # Definir el título del documento
    worksheet.merge_cells('A1:G1')
    title_cell = worksheet.cell(row=1, column=1)
    title_cell.value = "Programacion Menu Usuarios"
    title_cell.font = Font(size=20, bold=True)
    title_cell.alignment = Alignment(horizontal="center")

    # Definir los encabezados de la tabla
    headers = ['Usuario','Rut', 'Nombre Menu', 'Fecha Servicio', 'Cantidad Almuerzo','Opciones Menu','Fecha Selección']
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=2, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    # Añadir los datos de las programaciones
    for row_num, prog in enumerate(programaciones, 3):
        worksheet.cell(row=row_num, column=1).value = str(prog.usuario)
        worksheet.cell(row=row_num, column=2).value = str(prog.usuario.rut)
        worksheet.cell(row=row_num, column=3).value = prog.nom_menu
        worksheet.cell(row=row_num, column=4).value = prog.fecha_servicio.strftime('%Y-%m-%d')
        worksheet.cell(row=row_num, column=5).value = prog.cantidad_almuerzo
        worksheet.cell(row=row_num, column=6).value = prog._syncing
        worksheet.cell(row=row_num, column=7).value = prog.fecha_seleccion.strftime('%Y-%m-%d') if prog.fecha_seleccion else ''

    # Ajustar el ancho de las columnas
    for col_num, column in enumerate(worksheet.columns, 1):
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            if isinstance(cell, openpyxl.cell.cell.Cell):  # Verificar si la celda es del tipo correcto
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
        adjusted_width = (max_length + 2)
        column_letter = get_column_letter(col_num)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # Guardar el archivo Excel en un buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    # Escribir el contenido del buffer en la respuesta HTTP
    response.write(buffer.getvalue())
    buffer.close()

    return response